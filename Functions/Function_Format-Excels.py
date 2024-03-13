# ---> Importing used modules
import io
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
# --->


# ---> Function Excels
def formatExcels(df_data_all, namer):

    # -------------------------------- #
    # Define color usage:

    white = 'FFFFFFFF'
    black = 'FF000000'
    low_grey = 'FFEAECEE'

    # -------------------------------- #

    # Define general items:
    iten = {'font': 'Calibri', 'size': 10, 'type': 'solid', 'style': 'thin'}

    # -------------------------------- #
    # Define usage variables:

    # Customizations applied to all elements
    font_body_level = Font(name=iten['font'], size=iten['size'], bold=False, color=black)
    alignment_body_level = Alignment(wrap_text=False, indent=1)
    border = Border(left=Side(border_style=iten['style'], color=black), right=Side(border_style=iten['style'], color=black), top=Side(border_style=iten['style'], color=black), bottom=Side(border_style=iten['style'], color=black), diagonal=Side(border_style=iten['style'], color=black), diagonal_direction=0, outline=Side(border_style=iten['style'], color=black), vertical=Side(border_style=iten['style'], color=black), horizontal=Side(border_style=iten['style'], color=black))

    # Customizations applied to specific elements

    font_first_row = Font(name=iten['font'], size=iten['size'], bold=True, color=white)
    fill_first_row = PatternFill(fill_type=iten['type'], start_color=black)
    fill_body_level1 = PatternFill(fill_type=iten['type'], start_color=low_grey)
    fill_body_level2 = PatternFill(fill_type=iten['type'], start_color=low_grey)

    # -------------------------------- #

    # Read files with the respective variable to be treated
    d_data_all = df_data_all.to_dict('records')

    #################################### Part 2: Structuring ->

    # Create a Workbook object
    wb = Workbook()

    # Create a worksheet
    ws = wb.active

    # -------------------------------- #
    # Add values in tabs:

    # Create a list for future comparison
    vut = []

    # Classify sheet name
    ws.title = namer

    # Add the data from d_combined to the worksheet
    header = list(d_data_all[0].keys())
    ws.append(header)
    for item in d_data_all:
        values = list(item.values())
        values = [','.join(val) if isinstance(val, list) else val for val in values]
        ws.append(values)
    vut.append({'a': ws, 'b': d_data_all})

    # Perform adjustments to the created lists
    worksheets = [ws]
    for sheet in worksheets:
        sheet.freeze_panes = "C2"  # Keep the first row always visible along with the first two columns
        if sheet != ws:
            sheet.protection.sheet = True

    # -------------------------------- #

    # Perform a for loop to add parameters to each information tab
    for vut_ in vut:

        # Apply font_body_level to entire list
        for row in vut_['a'].iter_rows(min_row=1, max_row=len(vut_['b'])+1):
            for cell in row:
                cell.font = font_body_level
                cell.alignment = alignment_body_level
                cell.border = border

        # Apply fill_body_level to entire list
        for i, row in enumerate(vut_['a'].iter_rows(min_row=2), start=2):
            fill = fill_body_level1 if i % 2 == 0 else fill_body_level2
            for cell in row:
                cell.fill = fill

        # Apply first_row variables to the first row
        for coluna in range(1, vut_['a'].max_column + 1):
            cell = vut_['a'].cell(row=1, column=coluna)
            cell.value = cell.value.upper()
            cell.font = font_first_row
            cell.fill = fill_first_row

        # Apply lock and activate auto filter
        vut_['a'].auto_filter.ref = vut_['a'].dimensions
        # vut_['a'].protection.sheet = True

        # ----------------->

        # Module ready to auto adjust field width
        column_widths = []

        # Iterate through the desired columns
        for column in range(1, vut_['a'].max_column + 1):
            column_letter = get_column_letter(column)
            max_length = 0

            # Iterate through the cells of the column and get the maximum content length
            for cell in vut_['a'][column_letter]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except TypeError:
                    pass

            # Add a margin for the maximum length (optional)
            column_width = max_length + 2
            if column_width > 40:
                column_width = 40

            # Store the estimated column width
            column_widths.append(column_width)

        # Set column widths based on estimated sizes
        for column, column_width in enumerate(column_widths, start=1):
            column_letter = get_column_letter(column)
            vut_['a'].column_dimensions[column_letter].width = column_width

        # ----------------->

    # Final saving
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0) 

    return buffer.getvalue()
# --->
